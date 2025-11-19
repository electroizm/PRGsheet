"""
Fiyat Mikro Sistemi - Service Account Versiyonu
Merkezi config ve Service Account ile güvenli fiyat karşılaştırma

Özellikler:
- Service Account ile güvenli Google Sheets erişimi
- Merkezi config yönetimi (PRGsheet)
- Tüm hassas bilgiler ve dosya yolları PRGsheet'te saklanır
- Sessiz çalışma (sadece log dosyasına yazar)
"""

import sys
import os

# Parent directory'yi Python path'e ekle (central_config için)
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

import pandas as pd
import glob
import logging
from pathlib import Path
from typing import List, Dict, Optional

# Merkezi config manager'ı import et
from central_config import CentralConfigManager

# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

# Log dosyasina yaz (konsol yok)
# PyInstaller ile freeze edildiginde dosya yollarini duzelt
if getattr(sys, 'frozen', False):
    base_dir = Path(sys.executable).parent
else:
    base_dir = Path(__file__).parent

log_dir = base_dir / 'logs'
log_dir.mkdir(exist_ok=True)
log_file = log_dir / 'fiyat_mikro.log'

logging.basicConfig(
    level=logging.ERROR,  # Sadece hatalar
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION - Service Account ve Merkezi Config
# ============================================================================

class FiyatMikroConfig:
    """
    Service Account ve merkezi config ile yapılandırma

    Artık:
    - Service account credentials kodda YOK
    - Dosya yolları environment variable'da YOK
    - Ayarlar PRGsheets'ten çekiliyor
    """

    def __init__(self):
        try:
            # Merkezi config manager oluştur (Service Account otomatik başlar)
            self.config_manager = CentralConfigManager()

            # PRGsheets'ten ayarları yükle
            self.settings = self.config_manager.get_settings()

            logger.info("Config yüklendi")

        except Exception as e:
            logger.error(f"Config yükleme hatası: {e}")
            raise

    @property
    def sap_toptan_dir(self) -> str:
        """SAP TOPTAN dizini"""
        path = self.settings.get('SAP_TOPTAN_DIR')
        if not path:
            raise ValueError(
                "PRGsheet -> Ayar sayfasında SAP_TOPTAN_DIR ayarı eksik!\n"
                "Lütfen bu ayarı Global olarak ekleyin.\n"
                "Örnek değer: D:\\GoogleDrive\\Fiyat\\SAP\\TOPTAN"
            )
        return path

    @property
    def directories(self) -> List[str]:
        """Analiz edilecek dizinler listesi"""
        return [self.sap_toptan_dir]

# ============================================================================
# CSV READER
# ============================================================================

def read_utf16_csv(file_path: str) -> Optional[pd.DataFrame]:
    """
    UTF-16 kodlamalı CSV dosyasını okur

    Args:
        file_path: CSV dosya yolu

    Returns:
        DataFrame veya None
    """
    try:
        df = pd.read_csv(file_path, encoding='utf-16', sep='\t')
        return df
    except:
        try:
            df = pd.read_csv(file_path, encoding='utf-16le', sep='\t')
            return df
        except:
            try:
                df = pd.read_csv(file_path, encoding='utf-16', sep=';')
                return df
            except Exception as e:
                logger.error(f"CSV okuma hatası ({file_path}): {e}")
                return None

# ============================================================================
# FILE MANAGER
# ============================================================================

class FileManager:
    """CSV dosya yönetimi"""

    def __init__(self, directories: List[str]):
        self.directories = directories

    def get_all_csv_files(self) -> List[Dict[str, str]]:
        """
        Tüm dizinlerdeki CSV dosyalarını bulur

        Returns:
            CSV dosya bilgileri listesi
        """
        all_files = []
        for directory in self.directories:
            if os.path.exists(directory):
                pattern = os.path.join(directory, "*.csv")
                files = glob.glob(pattern)
                for file in files:
                    all_files.append({
                        'path': file,
                        'directory': directory,
                        'filename': os.path.basename(file)
                    })
            else:
                logger.warning(f"Dizin bulunamadı: {directory}")

        logger.info(f"{len(all_files)} CSV dosyası bulundu")
        return all_files

# ============================================================================
# DATA PROCESSOR
# ============================================================================

class DataProcessor:
    """Veri işleme ve filtreleme"""

    def __init__(self, file_manager: FileManager):
        self.file_manager = file_manager

    def filter_and_combine_data(self) -> Optional[str]:
        """
        Satır verisi 3 ile başlayan ve 9 karakterden uzun satırları birleştirir

        Returns:
            Çıkış dosyası yolu veya None
        """
        csv_files = self.file_manager.get_all_csv_files()

        if not csv_files:
            logger.error("Filtrelenecek CSV dosyaları bulunamadı!")
            return None

        filtered_toptan = []

        for file_info in csv_files:
            file_path = file_info['path']
            directory = file_info['directory']
            filename = file_info['filename']

            try:
                df = read_utf16_csv(file_path)
                if df is None:
                    continue

                # SAP Kodu sütununu kontrol et
                sap_col = 'SAP Kodu' if 'SAP Kodu' in df.columns else 'Kalem numarası'
                if sap_col in df.columns:
                    # 3 ile başlayan ve 9 karakterden uzun olanları filtrele
                    mask = df[sap_col].astype(str).str.startswith('3') & \
                           (df[sap_col].astype(str).str.len() > 9)

                    filtered_df = df[mask].copy()

                    if len(filtered_df) > 0:
                        filtered_df['Kaynak_Dosya'] = filename
                        filtered_df['Tam_Yol'] = file_path

                        if 'TOPTAN' in directory:
                            filtered_toptan.append(filtered_df)

            except Exception as e:
                logger.error(f"Dosya işleme hatası ({filename}): {e}")

        if not filtered_toptan:
            logger.warning("Filtreleme kriterlerine uyan satır bulunamadı!")
            return None

        # Excel dosyası oluştur
        output_file = os.path.join(
            self.file_manager.directories[0],
            'Filtrelenmis_Veriler_3_ile_baslayan.xlsx'
        )

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            toptan_combined = pd.concat(filtered_toptan, ignore_index=True)
            toptan_combined.to_excel(writer, sheet_name='TOPTAN', index=False)

        # Excel dosyasını yeniden aç ve başlık satırlarını düzenle
        from openpyxl import load_workbook

        wb = load_workbook(output_file)

        if 'TOPTAN' in wb.sheetnames:
            ws = wb['TOPTAN']

            # İlk satırın A1 hücresini sil
            ws['A1'] = None

            # İlk satırdaki tüm hücreleri 1 sütun sola kaydır
            max_col = ws.max_column
            for col in range(1, max_col):
                if col < max_col:
                    ws.cell(row=1, column=col).value = ws.cell(row=1, column=col+1).value

            # Son sütunu temizle
            ws.cell(row=1, column=max_col).value = None

        wb.save(output_file)
        wb.close()

        logger.info(f"Filtrelenmiş veri dosyası oluşturuldu: {output_file}")
        return output_file

    def create_price_comparison(self, filter_file: str) -> Optional[pd.DataFrame]:
        """
        TOPTAN fiyatlarını içeren karşılaştırma tablosu oluşturur

        Args:
            filter_file: Filtrelenmiş veri dosyası yolu

        Returns:
            Fiyat karşılaştırma DataFrame veya None
        """
        if not os.path.exists(filter_file):
            logger.error(f"Filtrelenmiş veri dosyası bulunamadı: {filter_file}")
            return None

        # TOPTAN verilerini oku
        try:
            toptan_df = pd.read_excel(filter_file, sheet_name='TOPTAN')
        except Exception as e:
            logger.error(f"TOPTAN sheet'i okunamadı: {e}")
            return None

        # TOPTAN verilerinden gerekli sütunları seç
        toptan_cols = ['SAP Kodu', 'Malzeme Adı', 'Fiyat', 'TOPTAN']

        # Eski sütun adlarını yeni adlarla eşleştir
        column_mapping = {
            'Ürün tanıtıcısı': 'SAP Kodu',
            'Tanım': 'Malzeme Adı',
            'Tutar': 'TOPTAN'
        }

        # Sütun adlarını güncelle
        toptan_df.rename(columns=column_mapping, inplace=True)

        missing_toptan_cols = [col for col in toptan_cols if col not in toptan_df.columns]
        if missing_toptan_cols:
            logger.error(f"TOPTAN'da eksik sütunlar: {missing_toptan_cols}")
            return None

        merged_df = toptan_df[toptan_cols].copy()

        # Boşlukları temizle ve integer'a çevir
        numeric_cols = ['Fiyat', 'TOPTAN']

        for col in numeric_cols:
            if col in merged_df.columns:
                try:
                    # String olan değerlerin başındaki ve sonundaki boşlukları temizle
                    merged_df[col] = merged_df[col].astype(str).str.strip()

                    # Virgüllü sayıları doğru şekilde integer'a çevir
                    merged_df[col] = merged_df[col].str.replace('.', '').str.replace(',', '.')
                    merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce')

                    # Float'ları integer'a çevir (NaN değerleri korunur)
                    merged_df[col] = merged_df[col].round().astype('Int64')

                except Exception as e:
                    logger.error(f"{col} sütunu işlenirken hata: {e}")

        # Sonucu Excel'e kaydet - çalışma dizininde
        output_file = os.path.join(os.getcwd(), 'Fiyat_Mikro.xlsx')
        merged_df.to_excel(output_file, index=False, engine='openpyxl')

        logger.info(f"Fiyat karşılaştırma dosyası oluşturuldu: {output_file}")
        return merged_df

    def cleanup_temp_files(self) -> None:
        """Geçici Excel dosyalarını sil"""
        files_to_delete = [
            os.path.join(self.file_manager.directories[0], 'tum_veriler_birlestirilmis.xlsx'),
            os.path.join(self.file_manager.directories[0], 'SAP_Verileri_Ayrilmis.xlsx'),
            os.path.join(self.file_manager.directories[0], 'Filtrelenmis_Veriler_3_ile_baslayan.xlsx')
        ]

        for file_path in files_to_delete:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"Geçici dosya silindi: {os.path.basename(file_path)}")
            except Exception as e:
                logger.error(f"Dosya silinemedi ({os.path.basename(file_path)}): {e}")

# ============================================================================
# GOOGLE SHEETS MANAGER - Service Account
# ============================================================================

class GoogleSheetsManager:
    """
    Service Account kullanan Google Sheets yöneticisi

    Artık:
    - Service account YOK
    - Service Account token kullanılıyor
    """

    def __init__(self, config_manager: CentralConfigManager):
        self.config_manager = config_manager
        self.gc = config_manager.gc  # Service Account ile yetkilendirilmiş client

    def upload_to_google_sheets(self, df: pd.DataFrame) -> None:
        """
        Fiyat karşılaştırma verilerini Google Sheets'e yükler

        Args:
            df: Yüklenecek DataFrame
        """
        try:
            # Sadece gerekli sütunları seç
            required_columns = ['SAP Kodu', 'Malzeme Adı', 'TOPTAN']

            # Sütun adlarını kontrol et ve düzelt
            upload_df = df.copy()
            if 'Malzeme Ad�' in upload_df.columns:
                upload_df.rename(columns={'Malzeme Ad�': 'Malzeme Adı'}, inplace=True)

            # Gerekli sütunları kontrol et
            missing_columns = [col for col in required_columns if col not in upload_df.columns]
            if missing_columns:
                logger.error(f"Eksik sütunlar: {missing_columns}")
                return

            # Sadece gerekli sütunları seç
            upload_df = upload_df[required_columns].copy()

            # PRGsheet'i doğrudan aç
            prgsheet = self.gc.open_by_key(self.config_manager.MASTER_SPREADSHEET_ID)

            # Fiyat_Mikro sayfasını kontrol et ve oluştur
            try:
                worksheet = prgsheet.worksheet('Fiyat_Mikro')
                worksheet.clear()
            except:
                worksheet = prgsheet.add_worksheet(title='Fiyat_Mikro', rows=2000, cols=10)

            # Verileri yükle
            if not upload_df.empty:
                values = [upload_df.columns.values.tolist()] + upload_df.values.tolist()
                worksheet.update(values, value_input_option='USER_ENTERED')
                logger.info(f"Google Sheets'e {len(upload_df)} satır yüklendi")
            else:
                logger.warning("Yüklenecek veri bulunamadı")

        except Exception as e:
            logger.error(f"Google Sheets upload hatası: {e}")
            raise

# ============================================================================
# FIYAT MIKRO ANALYZER
# ============================================================================

class FiyatMikroAnalyzer:
    """Fiyat Mikro ana sınıfı"""

    def __init__(self, config: FiyatMikroConfig):
        self.config = config
        self.file_manager = FileManager(config.directories)
        self.data_processor = DataProcessor(self.file_manager)
        self.sheets_manager = GoogleSheetsManager(config.config_manager)

    def run_analysis(self) -> None:
        """Ana fiyat analizi workflow'u çalıştır"""
        try:
            logger.info("Fiyat Mikro analizi başlatılıyor...")

            # 1. CSV dosyalarını bul
            csv_files = self.file_manager.get_all_csv_files()
            if not csv_files:
                logger.error("CSV dosyası bulunamadı!")
                return

            # 2. Verileri filtrele ve birleştir
            filter_file = self.data_processor.filter_and_combine_data()
            if not filter_file:
                logger.error("Filtreleme işlemi başarısız!")
                return

            # 3. Fiyat karşılaştırma tablosu oluştur
            price_df = self.data_processor.create_price_comparison(filter_file)
            if price_df is None or price_df.empty:
                logger.error("Fiyat karşılaştırma verisi oluşturulamadı!")
                return

            # 4. Google Sheets'e yükle (Service Account ile)
            self.sheets_manager.upload_to_google_sheets(price_df)

            # 5. Geçici dosyaları temizle
            self.data_processor.cleanup_temp_files()

            logger.info("[OK] Fiyat Mikro analizi tamamlandı!")

        except Exception as e:
            logger.error(f"[HATA] Fiyat Mikro analizi hatası: {e}")
            raise

# ============================================================================
# MAIN
# ============================================================================

def run_fiyat_mikro_analysis() -> None:
    """Ana fonksiyon - Sessiz mod (log dosyasına yazar)"""
    try:
        # Config oluştur (Service Account otomatik başlar)
        config = FiyatMikroConfig()

        # Analyzer oluştur
        analyzer = FiyatMikroAnalyzer(config)

        # Analiz çalıştır
        analyzer.run_analysis()

    except Exception as e:
        logger.error(f"Uygulama hatası: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    run_fiyat_mikro_analysis()
